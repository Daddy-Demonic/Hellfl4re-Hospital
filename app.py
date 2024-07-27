import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import re
from pathlib import Path
import random

# Define file path
desktop = Path.home() / "Desktop"
file_name = desktop / "hellfl4re_hospital.xlsx"

# Initialize the workbook
if file_name.exists():
    wb = load_workbook(file_name)
    sheet = wb.active
else:
    wb = Workbook()
    sheet = wb.active
    sheet.append(["ID", "Name", "Age", "Gender", "Address", "Contact", "Disease", "Medication", "Treating Doctor", "Guardian", "Aadhar Number"])

def save_workbook():
    wb.save(file_name)

def is_valid_name(name):
    return bool(re.match("^[a-zA-Z\n]+$", name))

def is_valid_age(age):
    return age.isdigit() and 0 < int(age) < 120

def is_valid_contact(contact):
    return bool(re.match("^\n{10}$", contact))

def is_valid_aadhar(aadhar):
    return bool(re.match("^\n{12}$", aadhar))

def generate_otp():
    return random.randint(100000, 999999)

def send_otp_verification():
    otp = generate_otp()
    # Mock sending OTP
    messagebox.showinfo("Verification", f"Your OTP is: {otp}")
    return otp

def validate_entries():
    name = name_entry.get()
    age = age_entry.get()
    gender = gender_var.get()
    address = address_entry.get()
    contact = contact_entry.get()
    disease = disease_entry.get()
    medication = medication_entry.get()
    treating_doctor = treating_doctor_entry.get()
    guardian = guardian_entry.get()
    aadhar = aadhar_entry.get()

    if not name or not age or not gender or not address or not contact or not disease or not medication or not treating_doctor or not guardian or not aadhar:
        messagebox.showerror("Error", "All fields are required.")
        return False

    if not is_valid_name(name):
        messagebox.showerror("Error", "Name must contain only alphabets.")
        return False

    if not is_valid_age(age):
        messagebox.showerror("Error", "Age must be a valid number between 1 and 119.")
        return False

    if not is_valid_contact(contact):
        messagebox.showerror("Error", "Contact must be a 10-digit number.")
        return False

    if not is_valid_aadhar(aadhar):
        messagebox.showerror("Error", "Aadhar number must be a 12-digit number.")
        return False

    return True

def add_patient():
    if not validate_entries():
        return

    otp = send_otp_verification()
    otp_entry = tk.Entry(root)
    otp_entry.grid(row=3, column=3, padx=10, pady=10)

    def verify_otp():
        entered_otp = otp_entry.get()
        if int(entered_otp) == otp:
            name = name_entry.get()
            age = age_entry.get()
            gender = gender_var.get()
            address = address_entry.get()
            contact = contact_entry.get()
            disease = disease_entry.get()
            medication = medication_entry.get()
            treating_doctor = treating_doctor_entry.get()
            guardian = guardian_entry.get()
            aadhar = aadhar_entry.get()

            next_id = sheet.max_row
            sheet.append([next_id, name, age, gender, address, contact, disease, medication, treating_doctor, guardian, aadhar])
            save_workbook()
            messagebox.showinfo("Success", "Patient added successfully!")
            clear_entries()
            view_patients()
            otp_entry.grid_forget()
            otp_label.grid_forget()
            verify_button.grid_forget()
        else:
            messagebox.showerror("Error", "Invalid OTP.")

    otp_label = tk.Label(root, text="Enter OTP:")
    otp_label.grid(row=3, column=2, padx=10, pady=10)

    verify_button = tk.Button(root, text="Verify OTP", command=verify_otp)
    verify_button.grid(row=4, column=2, padx=10, pady=10)

def view_patients():
    for row in patients_list.get_children():
        patients_list.delete(row)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        patients_list.insert('', 'end', values=row)

def clear_entries():
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    gender_var.set("Male")
    address_entry.delete(0, tk.END)
    contact_entry.delete(0, tk.END)
    disease_entry.delete(0, tk.END)
    medication_entry.delete(0, tk.END)
    treating_doctor_entry.delete(0, tk.END)
    guardian_entry.delete(0, tk.END)
    aadhar_entry.delete(0, tk.END)

def get_selected_patient(event):
    selected = patients_list.selection()
    if selected:
        values = patients_list.item(selected[0], 'values')
        clear_entries()
        name_entry.insert(0, values[1])
        age_entry.insert(0, values[2])
        gender_var.set(values[3])
        address_entry.insert(0, values[4])
        contact_entry.insert(0, values[5])
        disease_entry.insert(0, values[6])
        medication_entry.insert(0, values[7])
        treating_doctor_entry.insert(0, values[8])
        guardian_entry.insert(0, values[9])
        aadhar_entry.insert(0, values[10])

def update_patient():
    selected = patients_list.selection()
    if selected:
        if not validate_entries():
            return

        patient_id = patients_list.item(selected[0], 'values')[0]
        name = name_entry.get()
        age = age_entry.get()
        gender = gender_var.get()
        address = address_entry.get()
        contact = contact_entry.get()
        disease = disease_entry.get()
        medication = medication_entry.get()
        treating_doctor = treating_doctor_entry.get()
        guardian = guardian_entry.get()
        aadhar = aadhar_entry.get()

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == int(patient_id):
                row[1].value = name
                row[2].value = age
                row[3].value = gender
                row[4].value = address
                row[5].value = contact
                row[6].value = disease
                row[7].value = medication
                row[8].value = treating_doctor
                row[9].value = guardian
                row[10].value = aadhar
                save_workbook()
                messagebox.showinfo("Success", "Patient updated successfully!")
                clear_entries()
                view_patients()
                return
    else:
        messagebox.showerror("Error", "No patient selected.")

def delete_patient():
    selected = patients_list.selection()
    if selected:
        patient_id = patients_list.item(selected[0], 'values')[0]
        
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == int(patient_id):
                sheet.delete_rows(row[0].row)
                save_workbook()
                messagebox.showinfo("Success", "Patient deleted successfully!")
                clear_entries()
                view_patients()
                return
    else:
        messagebox.showerror("Error", "No patient selected.")

# Create main window
root = tk.Tk()
root.title("Hellfl4re Hospital Management System")

# Prevent resizing
root.geometry("1150x710")
root.resizable(False, False)

# Patient form
tk.Label(root, text="Name:").grid(row=0, column=0, padx=10, pady=10)
name_entry = tk.Entry(root)
name_entry.grid(row=0, column=1, padx=10, pady=10)

tk.Label(root, text="Age:").grid(row=1, column=0, padx=10, pady=10)
age_entry = tk.Entry(root)
age_entry.grid(row=1, column=1, padx=10, pady=10)

tk.Label(root, text="Gender:").grid(row=2, column=0, padx=10, pady=10)
gender_var = tk.StringVar(value="Male")
tk.Radiobutton(root, text="Male", variable=gender_var, value="Male").grid(row=2, column=1, padx=10, pady=10, sticky=tk.W)
tk.Radiobutton(root, text="Female", variable=gender_var, value="Female").grid(row=2, column=2, padx=10, pady=10, sticky=tk.W)

tk.Label(root, text="Address:").grid(row=3, column=0, padx=10, pady=10)
address_entry = tk.Entry(root)
address_entry.grid(row=3, column=1, padx=10, pady=10)

tk.Label(root, text="Contact:").grid(row=4, column=0, padx=10, pady=10)
contact_entry = tk.Entry(root)
contact_entry.grid(row=4, column=1, padx=10, pady=10)

tk.Label(root, text="Disease:").grid(row=5, column=0, padx=10, pady=10)
disease_entry = tk.Entry(root)
disease_entry.grid(row=5, column=1, padx=10, pady=10)

tk.Label(root, text="Medication:").grid(row=6, column=0, padx=10, pady=10)
medication_entry = tk.Entry(root)
medication_entry.grid(row=6, column=1, padx=10, pady=10)

tk.Label(root, text="Treating Doctor:").grid(row=7, column=0, padx=10, pady=10)
treating_doctor_entry = tk.Entry(root)
treating_doctor_entry.grid(row=7, column=1, padx=10, pady=10)

tk.Label(root, text="Guardian:").grid(row=8, column=0, padx=10, pady=10)
guardian_entry = tk.Entry(root)
guardian_entry.grid(row=8, column=1, padx=10, pady=10)

tk.Label(root, text="Aadhar Number:").grid(row=9, column=0, padx=10, pady=10)
aadhar_entry = tk.Entry(root)
aadhar_entry.grid(row=9, column=1, padx=10, pady=10)

# Buttons for CRUD operations
tk.Button(root, text="Add Patient", command=add_patient).grid(row=10, column=0, columnspan=2, pady=10)
tk.Button(root, text="Update Patient", command=update_patient).grid(row=10, column=1, columnspan=2, pady=10)
tk.Button(root, text="Delete Patient", command=delete_patient).grid(row=10, column=2, columnspan=2, pady=10)

# Patient list
cols = ("ID", "Name", "Age", "Gender", "Address", "Contact", "Disease", "Medication", "Treating Doctor", "Guardian", "Aadhar Number")
patients_list = ttk.Treeview(root, columns=cols, show='headings')
for col in cols:
    patients_list.heading(col, text=col)
    patients_list.column(col, width=100)
patients_list.grid(row=11, column=0, columnspan=4, padx=10, pady=10)

patients_list.bind('<<TreeviewSelect>>', get_selected_patient)

# Start the Tkinter main loop
view_patients()
root.mainloop()
